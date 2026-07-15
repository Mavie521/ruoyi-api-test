"""
一次性初始化 Excel 模板，之后不再用。
Excel 是数据源，数据归 Excel，代码归代码。
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

headers = ["id","title","feature","story","method","path","headers","params","json","data","check","expected","sql_check","sql_expected","jsonExData","sqlExData","marker","is_true","remark"]
hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = hf
    c.alignment = Alignment(horizontal="center")

for i, w in enumerate([10,30,12,12,8,25,35,25,55,25,25,15,40,15,35,35,6,8,20], 1):
    ws.column_dimensions[chr(64+i)].width = w

A = '{"Authorization":"Bearer {{TOKEN}}"}'

# 只放：登录、查询、新增（无环境依赖，无副作用）
cases = [
    ("L001","管理员登录","登录","正常","POST","/login","","",'{"username":"admin","password":"admin123"}',"","$.code","200","","",'{"TOKEN":"$.token"}',"","p0","TRUE"),
    ("L002","获取用户信息","登录","登录后","GET","/getInfo",A,"","","","$.code","200","","","","","p0","TRUE"),
    ("L003","获取菜单路由","登录","登录后","GET","/getRouters",A,"","","","$.code","200","","","","","p0","TRUE"),
    ("L004","错误密码","登录","异常","POST","/login","","",'{"username":"admin","password":"x"}',"","","$.code","500","","","","","p1","TRUE"),
    ("L005","不存在用户","登录","异常","POST","/login","","",'{"username":"nobody","password":"x"}',"","","$.code","500","","","","","p1","TRUE"),
    ("L006","空用户名","登录","异常","POST","/login","","",'{"username":"","password":"x"}',"","","$.code","500","","","","","p1","TRUE"),
    ("R001","查询角色列表","角色","查询","GET","/system/role/list",A,'{"pageNum":1,"pageSize":10}',"","","$.code","200","","","","","p0","TRUE"),
    ("R002","获取角色详情","角色","查询","GET","/system/role/1",A,"","","","","$.code","200","","","","","p0","TRUE"),
    ("R003","获取下拉选项","角色","查询","GET","/system/role/optionselect",A,"","","","","$.code","200","","","","","p0","TRUE"),
    ("R004","新增角色","角色","新增","POST","/system/role",A,"",'{"roleName":"excel_role","roleKey":"excel_key","roleSort":1,"status":"0","menuIds":[]}',"","","$.code","200","","","","","p0","TRUE"),
    ("U001","查询用户列表","用户","查询","GET","/system/user/list",A,'{"pageNum":1,"pageSize":10}',"","","$.code","200","","","","","p0","TRUE"),
    ("U002","获取管理员","用户","查询","GET","/system/user/1",A,"","","","","$.code","200","","","","","p0","TRUE"),
    ("U003","新增用户","用户","新增","POST","/system/user",A,"",'{"userName":"excel_user","nickName":"EXCEL","password":"123456","deptId":103,"email":"e@r.com","phonenumber":"13800000001","sex":"0","status":"0","postIds":[],"roleIds":[]}',"","","$.code","200","","","","","p0","TRUE"),
    ("U005","不存在用户","用户","异常","GET","/system/user/999999",A,"","","","","$.code","500","","","","","p1","TRUE"),
    ("R005","角色缺字段","角色","异常","POST","/system/role",A,"",'{"roleSort":1}',"","","$.code","500","","","","","p1","TRUE"),
]

for ri, case in enumerate(cases, 2):
    for ci, val in enumerate(case, 1):
        ws.cell(row=ri, column=ci, value=val)

wb.save("../data/test_cases.xlsx")
print(f"Excel模板已生成: {len(cases)}条用例")
print("之后编辑 data/test_cases.xlsx，不要改这个脚本")
