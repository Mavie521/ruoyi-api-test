"""Excel 测试用例格式化：说明页 + 颜色 + 下拉 + 冻结"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = openpyxl.load_workbook("data/test_cases.xlsx")

# ═══════════════════════════ 1. 使用说明 Sheet ═══════════════════════════
if "使用说明" in wb.sheetnames:
    del wb["使用说明"]
ws_help = wb.create_sheet("使用说明", 0)

help_data = [
    ["RuoYi API 测试用例编写指南", "", "", ""],
    ["", "", "", ""],
    ["一、整体结构", "", "", ""],
    ["  用例数据 Sheet — 一行一条用例，从上到下顺序执行", "", "", ""],
    ["  使用说明 Sheet — 本文档", "", "", ""],
    ["", "", "", ""],
    ["二、必填字段（缺一不可）", "", "", ""],
    ["  字段名", "作用", "填写示例", "填写规则"],
    ["  id", "用例编号", "L001", "唯一，建议格式：模块缩写+序号"],
    ["  method", "HTTP 方法", "get / post / put / delete", "小写字母，下拉选择"],
    ["  path", "接口路径", "/login", "以 / 开头，支持 {{变量}} 占位"],
    ["  is_true", "是否执行", "TRUE / FALSE", "TRUE=执行 FALSE=跳过，下拉选择"],
    ["  expected_json", "断言规则", '{"$.code": 200}', "JSON 格式，花括号包裹，允许多断言"],
    ["", "", "", ""],
    ["三、可选字段（按需填写，不填不影响执行）", "", "", ""],
    ["  feature", "Allure 模块名", "登录认证", "报告一级分组"],
    ["  story", "Allure 场景名", "正常登录", "报告二级分组"],
    ["  title", "用例标题", "管理员登录成功", "报告展示"],
    ["  headers", "请求头", '{"Content-Type": "application/json"}', "JSON 格式"],
    ["  params", "URL 参数", '{"pageNum": 1}', "GET 请求的查询字符串"],
    ["  json", "请求体", '{"username": "admin"}', "POST/PUT 的请求 Body"],
    ["  data", "表单数据", "", "文件上传场景用，一般留空"],
    ["  response_type", "响应类型", "binary / 空", "binary=文件下载，空=JSON"],
    ["  marker", "用例级别", "p0 / p1 / 空", "p0=冒烟 p1=回归 空=默认，下拉选择"],
    ["  remark", "备注", "任意文本", "不参与执行，给人看的说明"],
    ["", "", "", ""],
    ["四、高级字段（SQL 断言 + 变量提取）", "", "", ""],
    ["  sql_check", "SQL 查询", "SELECT status FROM sys_role WHERE role_key=''test'' ", "单引号注意转义"],
    ["  sql_expected", "SQL 预期值", "0", "与 sql_check 配对使用"],
    ["  jsonExData", "提取变量(JSONPath)", '{"ROLE_ID": "$.data.roleId"}', "从响应中提取变量给后续用例用"],
    ["  sqlExData", "提取变量(SQL)", '{"DEPT_ID": "SELECT dept_id FROM sys_dept LIMIT 1"}', "从数据库提取变量"],
    ["", "", "", ""],
    ["五、expected_json 断言写法", "", "", ""],
    ["  单断言:", '{"$.code": 200}', "只校验接口返回码", ""],
    ["  多断言:", '{"$.code": 200, "$.msg": "操作成功"}', "同时校验多个字段", ""],
    ["  复杂断言:", '{"$.code": 200, "$.data.rows[0].roleName": "管理员"}', "JSONPath 支持数组索引", ""],
    ["  类型转换:", '"200" -> 200, "true" -> True, "null" -> None', "字符串自动转 Python 类型", ""],
    ["  留空:", "不填 = 跳过断言", "不需要校验时直接空着", ""],
    ["", "", "", ""],
    ["六、变量传递（用例间数据流转）", "", "", ""],
    ["  内置变量: {{TIMESTAMP}} {{TS}} {{TOKEN}}", "无需手动创建，框架自动注入", "", ""],
    ["  提取变量: jsonExData 或 sqlExData 提取后，后续用例可以用 {{变量名}} 引用", "", "", ""],
    ["  示例: path=/system/user/{{USER_ID}}", "运行时自动替换为实际值", "", ""],
    ["  注意: 上一行提取的变量，下一行就能用；顺序执行，不要跳行引用", "", "", ""],
    ["", "", "", ""],
    ["七、运行命令", "", "", ""],
    ["  pytest testcases/ -v", "全量执行", "", ""],
    ["  pytest testcases/ -v -m p0", "只跑 P0 冒烟", "", ""],
    ["", "", "", ""],
    ["八、用例数据 Sheet 列颜色说明", "", "", ""],
    ["  红色表头 = 必填字段", "id / method / path / is_true / expected_json", "", ""],
    ["  绿色表头 = 可选字段", "feature / story / title / headers / params / json / data / remark", "", ""],
    ["  紫色表头 = SQL 断言", "sql_check / sql_expected / jsonExData / sqlExData", "", ""],
    ["  蓝色表头 = 文件下载", "expected_content_type / expected_content_min_bytes / expected_content_sha256", "", ""],
]

TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="C00000")
SECTION_FONT = Font(name="微软雅黑", bold=True, size=11, color="1F4E79")
SECTION_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
TABLE_HEADER = Font(name="微软雅黑", bold=True, size=10)
TABLE_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
NORMAL = Font(name="微软雅黑", size=10)
CODE = Font(name="Consolas", size=10)

for r, row_data in enumerate(help_data, 1):
    for c, val in enumerate(row_data, 1):
        cell = ws_help.cell(row=r, column=c, value=val)
        if r == 1:
            cell.font = TITLE_FONT
        elif val and isinstance(val, str) and val.startswith(("一、", "二、", "三、", "四、", "五、", "六、", "七、")):
            for cc in range(1, 5):
                ws_help.cell(row=r, column=cc).font = SECTION_FONT
                ws_help.cell(row=r, column=cc).fill = SECTION_FILL
        elif r in (8, 17, 28, 34):
            for cc in range(1, 5):
                ws_help.cell(row=r, column=cc).font = TABLE_HEADER
                ws_help.cell(row=r, column=cc).fill = TABLE_HEADER_FILL
        elif c == 1 and val and val.startswith("  "):
            cell.font = CODE if "{" in str(val) else NORMAL
        else:
            cell.font = NORMAL

ws_help.column_dimensions["A"].width = 18
ws_help.column_dimensions["B"].width = 22
ws_help.column_dimensions["C"].width = 60
ws_help.column_dimensions["D"].width = 40

# ═══════════════════════════ 2. 用例数据 Sheet 格式化 ═══════════════════════════
ws = wb["Sheet"]
# 清理之前可能残留的合并单元格
for mc in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(mc))

# 颜色分组
RED_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
GREEN_FILL = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
BLUE_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# 表头颜色：红=必填 绿=可选 紫=SQL 蓝=文件
required = {"id", "method", "path", "is_true", "expected_json"}
purple = {"sql_check", "sql_expected", "jsonExData", "sqlExData"}
blue = {"expected_content_type", "expected_content_min_bytes", "expected_content_sha256"}

for col_idx, cell in enumerate(ws[1], 1):  # 表头在第1行
    col_name = str(cell.value or "")
    cell.font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if col_name in required:
        cell.fill = RED_FILL
    elif col_name in purple:
        cell.fill = PURPLE_FILL
    elif col_name in blue:
        cell.fill = BLUE_FILL
    else:
        cell.fill = GREEN_FILL

# 冻结表头
ws.freeze_panes = "A2"

# 列宽
widths = {"A": 8, "B": 12, "C": 12, "D": 18, "E": 8, "F": 32, "G": 20, "H": 16,
          "I": 45, "J": 16, "K": 38, "L": 14, "M": 16, "N": 16, "O": 16,
          "P": 45, "Q": 16, "R": 26, "S": 26, "T": 8, "U": 8, "V": 8}
for letter, w in widths.items():
    ws.column_dimensions[letter].width = w

# 数据行对齐
for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(vertical="center")

# 下拉选择
dv1 = DataValidation(type="list", formula1='"get,post,put,delete"', allow_blank=True)
ws.add_data_validation(dv1)
dv1.add("E2:E200")

dv2 = DataValidation(type="list", formula1='"p0,p1,p2"', allow_blank=True)
ws.add_data_validation(dv2)
dv2.add("T2:T200")

dv3 = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
ws.add_data_validation(dv3)
dv3.add("U2:U200")

dv4 = DataValidation(type="list", formula1='"binary,"', allow_blank=True)
ws.add_data_validation(dv4)
dv4.add("L2:L200")

# is_true=FALSE → 灰底
ws.conditional_formatting.add("A2:V200", CellIsRule(
    operator="equal", formula=['"FALSE"'],
    fill=PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    font=Font(color="808080")
))

wb.save("data/test_cases.xlsx")
print("Done: 使用说明 + 颜色图例 + 下拉选择 + 冻结表头 + FALSE灰底")
