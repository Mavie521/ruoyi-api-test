"""生成Excel测试用例"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'

# 样式
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='center', horizontal='center')

# 第1行标题
ws.merge_cells('A1:O1')
ws['A1'] = '若依接口测试 - Excel数据驱动（is_true=True的用例才会执行）'
ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
ws['A1'].fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

# 表头
headers = ['id','feature','story','title','severity','method','path','headers','params',
           'json','data','check','expected','jsonExData','is_true']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = wrap_align
    cell.border = thin_border

# 测试数据
cases = [
    # 登录模块
    ['TC001','登录模块','登录成功','管理员登录成功','阻塞','post','/login',
     '{"Content-Type":"application/json"}','',
     '{"username":"admin","password":"admin123"}','',
     '$..msg','操作成功','{"TOKEN":"$..token"}','True'],

    ['TC002','登录模块','登录异常','错误密码登录失败','严重','post','/login',
     '{"Content-Type":"application/json"}','',
     '{"username":"admin","password":"wrong_pass"}','',
     '$..msg','用户不存在/密码错误','','True'],

    ['TC003','登录模块','登录异常','空用户名登录失败','一般','post','/login',
     '{"Content-Type":"application/json"}','',
     '{"username":"","password":"admin123"}','',
     '$..msg','用户不存在/密码错误','','True'],

    ['TC004','登录模块','登录后查询','获取当前用户信息','阻塞','get','/getInfo',
     '{"Authorization":"Bearer {{TOKEN}}"}','','','','$..code','200','','True'],

    ['TC005','登录模块','登录后查询','获取菜单路由','严重','get','/getRouters',
     '{"Authorization":"Bearer {{TOKEN}}"}','','','','$..code','200','','True'],

    # 真实用户管理
    ['TC006','用户管理(真实)','用户查询','查询用户列表','阻塞','get','/system/user/list',
     '{"Authorization":"Bearer {{TOKEN}}"}','{"pageNum":1,"pageSize":10}','','','$..code','200','','True'],

    ['TC007','用户管理(真实)','用户查询','获取管理员详情(id=1)','严重','get','/system/user/1',
     '{"Authorization":"Bearer {{TOKEN}}"}','','','','$..data.userName','admin','','True'],

    # 角色管理
    ['TC008','角色管理(真实)','角色查询','查询角色列表','阻塞','get','/system/role/list',
     '{"Authorization":"Bearer {{TOKEN}}"}','{"pageNum":1,"pageSize":10}','','','$..code','200','','True'],

    ['TC009','角色管理(真实)','角色查询','获取角色下拉选项','严重','get','/system/role/optionselect',
     '{"Authorization":"Bearer {{TOKEN}}"}','','','','$..code','200','','True'],

    ['TC010','角色管理(真实)','角色查询','获取超级管理员角色','严重','get','/system/role/1',
     '{"Authorization":"Bearer {{TOKEN}}"}','','','','$..data.roleName','超级管理员','','True'],

    # 部门管理
    ['TC011','部门管理(真实)','部门查询','获取部门列表','阻塞','get','/system/dept/list',
     '{"Authorization":"Bearer {{TOKEN}}"}','','','','$..code','200','','True'],

    ['TC012','部门管理(真实)','部门查询','获取部门详情(root)','一般','get','/system/dept/100',
     '{"Authorization":"Bearer {{TOKEN}}"}','','','','$..data.deptName','若依科技','','True'],

    # 字典管理
    ['TC013','字典管理(真实)','字典查询','获取字典类型列表','一般','get','/system/dict/type/list',
     '{"Authorization":"Bearer {{TOKEN}}"}','{"pageNum":1,"pageSize":10}','','','$..code','200','','True'],

    ['TC014','字典管理(真实)','字典查询','获取字典数据','一般','get','/system/dict/data/type/sys_normal_disable',
     '{"Authorization":"Bearer {{TOKEN}}"}','','','','$..code','200','','True'],

    # 岗位管理
    ['TC015','岗位管理(真实)','岗位查询','获取岗位列表','一般','get','/system/post/list',
     '{"Authorization":"Bearer {{TOKEN}}"}','{"pageNum":1,"pageSize":10}','','','$..code','200','','True'],

    # 系统监控
    ['TC016','系统监控','监控查询','获取在线用户列表','一般','get','/monitor/online/list',
     '{"Authorization":"Bearer {{TOKEN}}"}','{"pageNum":1,"pageSize":10}','','','$..code','200','','True'],

    ['TC017','系统监控','监控查询','获取操作日志列表','一般','get','/monitor/operlog/list',
     '{"Authorization":"Bearer {{TOKEN}}"}','{"pageNum":1,"pageSize":10}','','','$..code','200','','True'],

    ['TC018','系统监控','监控查询','获取登录日志列表','一般','get','/monitor/logininfor/list',
     '{"Authorization":"Bearer {{TOKEN}}"}','{"pageNum":1,"pageSize":10}','','','$..code','200','','True'],

    # 测试控制器
    ['TC019','用户管理(测试)','用户查询','获取测试用户列表','一般','get','/test/user/list',
     '{"Authorization":"Bearer {{TOKEN}}"}','','','','$..code','200','','True'],

    # 包含断言示例
    ['TC020','登录模块','包含断言','登录成功-响应包含操作成功','一般','post','/login',
     '{"Content-Type":"application/json"}','',
     '{"username":"admin","password":"admin123"}','',
     '','操作成功','','True'],
]

# 写入
for row_idx, case in enumerate(cases, 3):
    for col_idx, value in enumerate(case, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        if col_idx in (8,9,10,14):
            cell.font = Font(size=9, name='Consolas')

# 列宽
col_widths = [7,14,12,24,8,7,30,32,22,30,22,14,18,22,7]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 冻结表头
ws.freeze_panes = 'A3'

wb.save('D:/Code/claude_test01/ruoyi_api_test/data/test_cases.xlsx')
print(f'OK! 生成了 {len(cases)} 条测试用例')
